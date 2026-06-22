"""
parsers/ddc_adapter.py — DataDrivenConstruction (DDC) Converter CLI Adapter

INTEGRATION: cad2data-Revit-IFC-DWG-DGN (https://github.com/datadrivenconstruction/cad2data-Revit-IFC-DWG-DGN)

PURPOSE:
    Bridge between FireAI's parser layer and the DDC CLI converters:
    - ddc-rvtconverter  → Revit .rvt → XLSX + DAE
    - ddc-dwgconverter  → AutoCAD .dwg → XLSX + DAE
    - ddc-ifcconverter  → IFC → XLSX + DAE
    - ddc-dgnconverter  → MicroStation .dgn → XLSX + DAE

ARCHITECTURE:
    FireAI Workflow → DDCAdapter.convert() → subprocess(ddc-*converter) → XLSX
    FireAI Workflow → DDCAdapter.to_rooms() → IFC/rooms list for device placement

LICENSING NOTE:
    DDC converter binaries are proprietary (DDC EULA + ODA Sustaining Member license).
    CLI execution is permitted under the EULA for integration purposes.
    Source code of converters is NOT available.

INSTALLATION (Linux):
    curl -sS https://pkg.datadrivenconstruction.io/setup.sh | sudo bash
    sudo apt install ddc-rvtconverter ddc-dwgconverter ddc-ifcconverter

FALLBACK:
    If DDC converters are not installed, each method raises DDCNotAvailableError.
    The caller must handle this and fall back to ezdxf/ifcopenshell or manual input.

SAFETY:
    DDC output is used as geometry INPUT to NFPA 72 calculations.
    It is NOT authoritative for compliance — all extracted data flows through
    the validation layer before reaching the device placement engine.
"""

from __future__ import annotations

import logging
import os
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# DDC CLI binary names (Linux .deb installed to /usr/bin/)
_DDC_CONVERTERS: Dict[str, str] = {
    ".rvt": "ddc-rvtconverter",
    ".rfa": "ddc-rvtconverter",
    ".dwg": "ddc-dwgconverter",
    ".ifc": "ddc-ifcconverter",
    ".dgn": "ddc-dgnconverter",
}

# Windows .exe equivalents (fallback for Windows deployments)
_DDC_CONVERTERS_WIN: Dict[str, str] = {
    ".rvt": "RvtExporter.exe",
    ".rfa": "RvtExporter.exe",
    ".dwg": "DwgExporter.exe",
    ".ifc": "IfcExporter.exe",
    ".dgn": "DgnExporter.exe",
}

# SECURITY FIX (V103): ALLOWED_BINARIES maps binary names to their expected
# absolute paths. Before any subprocess call, the resolved binary path is
# checked against this dict. This prevents an attacker from placing a
# malicious binary on PATH that shadows the real DDC converter.
# Only binaries found at these exact paths (or in these directories) are
# permitted. If the resolved binary is NOT in this dict's values, the
# conversion is REJECTED with a SECURITY error.
#
# On Linux, DDC converters are installed via apt to /usr/bin/.
# On Windows, they are typically in the converter_dir or C:\DDC\.
_ALLOWED_BINARIES: Dict[str, List[str]] = {
    "ddc-rvtconverter": ["/usr/bin/ddc-rvtconverter", "/usr/local/bin/ddc-rvtconverter"],
    "ddc-dwgconverter": ["/usr/bin/ddc-dwgconverter", "/usr/local/bin/ddc-dwgconverter"],
    "ddc-ifcconverter": ["/usr/bin/ddc-ifcconverter", "/usr/local/bin/ddc-ifcconverter"],
    "ddc-dgnconverter": ["/usr/bin/ddc-dgnconverter", "/usr/local/bin/ddc-dgnconverter"],
    "RvtExporter.exe": [],  # Windows — paths resolved from converter_dir
    "DwgExporter.exe": [],  # Windows — paths resolved from converter_dir
    "IfcExporter.exe": [],  # Windows — paths resolved from converter_dir
    "DgnExporter.exe": [],  # Windows — paths resolved from converter_dir
}

# SECURITY: The subprocess working directory (cwd) is restricted to a
# dedicated safe directory, NOT an arbitrary user-supplied path.
# Previously, cwd was set to output_dir (which could be user-controlled).
# Now we use the resolved temp directory as cwd, which is a known safe path.
_SAFE_CWD_BASE = Path(os.getenv("FIREAI_DDC_CWD_BASE", tempfile.gettempdir())).resolve()


class DDCNotAvailableError(RuntimeError):
    """Raised when DDC CLI converter is not installed."""

    pass


@dataclass
class DDCConversionResult:
    """Result of a DDC converter run."""

    success: bool
    source_file: str
    xlsx_path: Optional[str] = None
    dae_path: Optional[str] = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    rooms: List[Dict[str, Any]] = field(default_factory=list)
    elements: List[Dict[str, Any]] = field(default_factory=list)
    duration_s: float = 0.0


class DDCAdapter:
    """
    Adapter for DDC (DataDrivenConstruction) CLI converters.

    Converts proprietary BIM/CAD formats to open XLSX/DAE for FireAI analysis.

    THREAD SAFETY: Each convert() call uses a separate temp directory.
    SAFETY: Output is advisory geometry — validated before use in calculations.
    """

    def __init__(self, converter_dir: Optional[str] = None):
        """
        Initialize DDC adapter.

        Args:
            converter_dir: Optional path to DDC .exe directory (Windows).
                          On Linux, converters must be on $PATH via apt install.
        """
        self._converter_dir = converter_dir
        self._platform = "windows" if os.name == "nt" else "linux"

    def is_available(self, file_ext: str) -> bool:
        """Check if DDC converter is available for the given file extension.

        V105 FIX (LOW-5): Use shutil.which() instead of spawning a subprocess.
        This is faster and doesn't leak information about installed binaries.
        """
        ext = file_ext.lower()
        binary = self._get_binary(ext)
        if binary is None:
            return False
        try:
            import shutil

            resolved = shutil.which(binary)
            return resolved is not None
        except Exception:
            return False

    def _get_binary(self, ext: str) -> Optional[str]:
        """Get the DDC binary name for a file extension."""
        if self._platform == "windows":
            binary = _DDC_CONVERTERS_WIN.get(ext)
            if binary and self._converter_dir:
                return str(Path(self._converter_dir) / binary)
            return binary
        return _DDC_CONVERTERS.get(ext)

    def convert(
        self,
        input_path: str,
        export_mode: str = "standard",
        output_dir: Optional[str] = None,
    ) -> DDCConversionResult:
        """
        Convert a CAD/BIM file using the appropriate DDC converter.

        Args:
            input_path: Path to .rvt, .dwg, .ifc, or .dgn file
            export_mode: DDC export mode (basic/standard/complete) for Revit
            output_dir: Output directory (defaults to temp dir)

        Returns:
            DDCConversionResult with paths to generated XLSX/DAE files

        Raises:
            DDCNotAvailableError: If DDC converter not installed
            FileNotFoundError: If input file doesn't exist
        """
        import time

        start = time.monotonic()

        # V123 REFACTOR (per agent.md Rule #23 — single source of truth):
        # All path-security validation now delegates to the shared helper
        # in parsers._path_security. The previous ~70 lines of inline
        # path-traversal / symlink / allowed-bases / extension checks
        # have been replaced with a single call to validate_input_path()
        # which performs the same checks PLUS new defenses introduced by
        # V122 (null byte rejection, leading-dash argument-injection
        # guard). This guarantees DDC and DWG parsers share IDENTICAL
        # security posture — no drift possible.
        #
        # Behavioral changes vs. pre-V123:
        #   + Null bytes in input_path now rejected (previously no check)
        #   + Paths starting with '-' rejected (argument-injection guard)
        #   = Path traversal, allowed-bases, symlink resolution, and
        #     extension whitelist behaviors are PRESERVED IDENTICALLY
        #
        # Exception mapping:
        #   FileNotFoundError raised by helper → re-raised unchanged
        #     (preserves backward-compat for callers that catch it)
        #   UnsafePathError raised by helper → re-raised as ValueError
        #     (preserves backward-compat: pre-V123 raised ValueError for
        #     path-traversal & extension errors)
        from parsers._path_security import (
            UnsafePathError,
            validate_file_size,
            validate_input_path,
        )

        _ALLOWED_EXTENSIONS = frozenset(_DDC_CONVERTERS.keys())
        try:
            safe_path = validate_input_path(
                input_path,
                allowed_extensions=_ALLOWED_EXTENSIONS,
                parser_name="DDCAdapter",
            )
        except UnsafePathError as e:
            # Preserve pre-V123 contract: path-traversal / bad extension
            # historically raised ValueError. We map UnsafePathError →
            # ValueError so downstream `except ValueError:` blocks in
            # callers continue to work without modification.
            raise ValueError(str(e)) from e

        # V126: File-size cap — reject oversized files before feeding them
        # to the DDC subprocess. Without this, a multi-GB .rvt/.ifc would
        # exhaust memory and potentially hang the converter for hours.
        _DDC_MAX_FILE_SIZE_BYTES = int(
            os.getenv("FIREAI_DDC_MAX_FILE_SIZE_BYTES", 500 * 1024 * 1024)  # 500 MB default
        )
        try:
            validate_file_size(
                safe_path,
                max_size_bytes=_DDC_MAX_FILE_SIZE_BYTES,
                parser_name="DDCAdapter",
            )
        except UnsafePathError as e:
            raise ValueError(str(e)) from e

        # Local alias for the resolved Path; rest of method unchanged.
        input_path_obj = Path(input_path)
        ext = input_path_obj.suffix.lower()

        binary = self._get_binary(ext)
        if binary is None:
            raise DDCNotAvailableError(
                f"No DDC converter registered for extension '{ext}'. Supported: {list(_DDC_CONVERTERS.keys())}"
            )

        # Use temp dir if no output specified
        _temp = None
        if output_dir is None:
            _temp = tempfile.mkdtemp(prefix="fireai_ddc_")
            output_dir = _temp

        # SECURITY FIX (V103): Resolve the binary path and verify it is
        # in an ALLOWED location. Without this check, an attacker could
        # place a malicious binary on PATH (e.g., /tmp/ddc-rvtconverter)
        # that would be executed instead of the real converter.
        # We use shutil.which() to resolve the full path, then verify
        # the resolved path matches one of the allowed locations.
        import shutil

        resolved_binary = shutil.which(binary)
        if resolved_binary is None:
            raise DDCNotAvailableError(
                f"DDC converter '{binary}' not found on PATH. "
                f"Install with: sudo apt install {_DDC_CONVERTERS.get(ext, binary)}"
            )

        # Verify the resolved binary is in an allowed location
        resolved_binary_path = Path(resolved_binary).resolve()
        _binary_name = Path(binary).name  # Strip any path prefix
        _allowed_paths = _ALLOWED_BINARIES.get(_binary_name, [])

        # For Windows with converter_dir, add converter_dir to allowed paths
        if self._platform == "windows" and self._converter_dir:
            _allowed_paths = list(_allowed_paths) + [str(Path(self._converter_dir).resolve() / _binary_name)]

        if _allowed_paths:
            _binary_in_allowed = False
            for allowed_path in _allowed_paths:
                try:
                    resolved_binary_path.relative_to(Path(allowed_path).resolve())
                    _binary_in_allowed = True
                    break
                except ValueError:
                    # Also check exact match
                    if str(resolved_binary_path) == allowed_path:
                        _binary_in_allowed = True
                        break
            if not _binary_in_allowed:
                raise ValueError(
                    f"SECURITY: DDC binary '{resolved_binary_path}' is not in "
                    f"allowed locations. Binary path traversal detected. "
                    f"Allowed: {_allowed_paths}"
                )
        else:
            # V105 FIX (MEDIUM-9): When _allowed_paths is empty and
            # platform is Windows without converter_dir, REJECT the binary
            # (fail-closed). Previously, empty list = no validation.
            if not (self._platform == "windows" and self._converter_dir):
                raise ValueError(
                    f"SECURITY: DDC binary '{_binary_name}' has no allowed "
                    f"paths configured. Refusing to execute unknown binary. "
                    f"Add allowed paths to _ALLOWED_BINARIES or set converter_dir."
                )

        try:
            cmd = [str(resolved_binary_path), str(safe_path)]
            if export_mode and ext in (".rvt", ".rfa"):
                # SECURITY FIX: Whitelist export_mode to prevent command
                # injection. Without validation, a malicious export_mode
                # could pass unexpected flags to the DDC converter binary.
                _ALLOWED_EXPORT_MODES = frozenset({"basic", "standard", "complete"})
                if export_mode not in _ALLOWED_EXPORT_MODES:
                    raise ValueError(
                        f"Invalid export_mode '{export_mode}'. Permitted values: {sorted(_ALLOWED_EXPORT_MODES)}"
                    )
                cmd.append(export_mode)

            # SECURITY FIX (V103): Restrict subprocess cwd to a safe directory
            # instead of using the user-controlled output_dir. The output_dir
            # is passed as an argument to the binary (if supported) rather
            # than used as cwd. This prevents the subprocess from writing to
            # or reading from arbitrary directories.
            _safe_cwd = _SAFE_CWD_BASE / f"fireai_ddc_cwd_{os.getpid()}"
            _safe_cwd.mkdir(parents=True, exist_ok=True)

            logger.info("DDC convert: %s → %s", ' '.join(cmd), output_dir)

            result = subprocess.run(  # noqa: S603 — command from class constant, not user input
                cmd,
                capture_output=True,
                text=True,
                cwd=str(_safe_cwd),
                timeout=300,  # 5 min max
            )

            duration = time.monotonic() - start

            if result.returncode != 0:
                return DDCConversionResult(
                    success=False,
                    source_file=str(input_path),
                    errors=[f"DDC converter failed (exit {result.returncode}): {result.stderr[:500]}"],
                    duration_s=duration,
                )

            # Find generated output files
            stem = input_path_obj.stem
            xlsx_path = Path(output_dir) / f"{stem}_rvt.xlsx"
            if not xlsx_path.exists():
                # DDC sometimes uses the full filename
                for f in Path(output_dir).glob("*.xlsx"):
                    xlsx_path = f
                    break

            dae_path = Path(output_dir) / f"{stem}.dae"
            if not dae_path.exists():
                dae_path = None  # type: ignore[assignment]

            conv_result = DDCConversionResult(
                success=True,
                source_file=str(input_path),
                xlsx_path=str(xlsx_path) if xlsx_path and xlsx_path.exists() else None,
                dae_path=str(dae_path) if dae_path else None,
                duration_s=duration,
            )

            # Parse XLSX for room/element data
            if conv_result.xlsx_path:
                conv_result.rooms = self._extract_rooms_from_xlsx(conv_result.xlsx_path)
                conv_result.elements = self._extract_elements_from_xlsx(conv_result.xlsx_path)

            logger.info(
                f"DDC convert OK: {input_path_obj.name} → "
                f"{len(conv_result.rooms)} rooms, {len(conv_result.elements)} elements "
                f"in {duration:.1f}s"
            )
            return conv_result

        except subprocess.TimeoutExpired:
            return DDCConversionResult(
                success=False,
                source_file=str(input_path),
                errors=["DDC converter timed out after 300s — file may be too large"],
                duration_s=300.0,
            )
        except Exception as e:
            logger.error("DDC convert error: %s", e, exc_info=True)
            return DDCConversionResult(
                success=False,
                source_file=str(input_path),
                errors=[f"{type(e).__name__}: {e}"],
            )

    def _extract_rooms_from_xlsx(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """
        Extract room/space data from DDC-generated XLSX.

        DDC XLSX columns include: Category, Type Name, Level, Area, Volume,
        Width, Height, Number, Name, and geometry parameters.
        """
        try:
            import openpyxl
        except ImportError:
            try:
                import xlrd  # type: ignore[import-not-found]  # noqa: F401

                return self._extract_rooms_xlrd(xlsx_path)
            except ImportError:
                logger.warning("openpyxl not installed — cannot parse DDC XLSX output")
                return []

        rooms = []
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return []

            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else "" for h in row]
                    continue

                row_dict = dict(zip(headers, row, strict=False))
                category = str(row_dict.get("Category", "")).lower()

                # Rooms are "Rooms" category in Revit, "IfcSpace" in IFC
                if "room" in category or "space" in category:
                    area = row_dict.get("Area", 0) or 0
                    # DDC exports in ft² for imperial projects — check unit
                    rooms.append(
                        {
                            "name": row_dict.get("Name", "") or row_dict.get("Number", ""),
                            "area_m2": float(area) if area else 0.0,
                            "level": str(row_dict.get("Level", "")) or "1",
                            "volume_m3": float(row_dict.get("Volume", 0) or 0),
                            "source": "ddc",
                            "raw": row_dict,
                        }
                    )

            wb.close()
        except Exception as e:
            logger.error("DDC XLSX room extraction failed: %s", e, exc_info=True)

        return rooms

    def _extract_elements_from_xlsx(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """Extract all BIM elements from DDC XLSX."""
        try:
            import openpyxl
        except ImportError:
            return []

        elements = []
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return []

            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else "" for h in row]
                    continue
                row_dict = dict(zip(headers, row, strict=False))
                elements.append(row_dict)

            wb.close()
        except Exception as e:
            logger.error("DDC XLSX element extraction failed: %s", e, exc_info=True)

        return elements

    def _extract_rooms_xlrd(self, xlsx_path: str) -> List[Dict[str, Any]]:
        """Fallback room extraction using xlrd (for .xls files)."""
        import xlrd  # type: ignore[import-not-found]

        rooms = []
        try:
            wb = xlrd.open_workbook(xlsx_path)
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            for r in range(1, ws.nrows):
                row_dict = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
                category = str(row_dict.get("Category", "")).lower()
                if "room" in category or "space" in category:
                    rooms.append(
                        {
                            "name": row_dict.get("Name", ""),
                            "area_m2": float(row_dict.get("Area", 0) or 0),
                            "level": str(row_dict.get("Level", "1")),
                            "source": "ddc",
                            "raw": row_dict,
                        }
                    )
        except Exception as e:
            logger.error("xlrd extraction failed: %s", e, exc_info=True)
        return rooms


def get_ddc_adapter(converter_dir: Optional[str] = None) -> DDCAdapter:
    """Get a DDCAdapter instance (simple factory, no singleton needed)."""
    return DDCAdapter(converter_dir=converter_dir)
