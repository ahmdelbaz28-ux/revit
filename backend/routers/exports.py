from typing import Dict, Optional, Tuple

from pydantic import BaseModel


def _generate_excel_export(project, devices, connections):
    """Generate Excel export content with multiple sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(  # NOSONAR — S8415
            status_code=503,
            detail={
                "success": False,
                "error": "Excel export unavailable: openpyxl package not installed",
                "install": "pip install openpyxl",
            },
        )

    wb = Workbook()

    # ── Sheet 1: Project metadata ──────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Project"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, val in enumerate(
        ["Field", "Value"], start=1
    ):
        cell = ws_proj.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
    proj_rows = [
        ("Project ID", project.get("id", "")),
        ("Project Name", project.get("name", "")),
        ("Author", project.get("author", "")),
        ("Exported At (UTC)", datetime.now(timezone.utc).isoformat()),
        ("Software", "FireAI / BAZSPARK v1.55.0 (V213)"),
        ("Device Count", len(devices)),
        ("Connection Count", len(connections)),
        ("Standard", "NFPA 72-2022"),
    ]
    for i, (k, v) in enumerate(proj_rows, start=2):
        ws_proj.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_proj.cell(row=i, column=2, value=str(v))
    ws_proj.column_dimensions["A"].width = 24
    ws_proj.column_dimensions["B"].width = 48

    # ── Sheet 2: Devices ───────────────────────────────────────────
    ws_dev = wb.create_sheet("Devices")
    dev_cols = [
        "ID", "Name", "Type", "Category",
        "X", "Y", "Z", "Rotation",
        "Voltage (V)", "Current (A)", "Load (W)",
        "Created At", "Updated At", "Properties",
    ]
    for col, name in enumerate(dev_cols, start=1):
        cell = ws_dev.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, d in enumerate(devices, start=2):
        props = d.get("properties") or {}
        props_str = json.dumps(props, ensure_ascii=False) if props else ""
        row_vals = [
            d.get("id", ""), d.get("name", ""),
            d.get("type", ""), d.get("category", ""),
            d.get("x", 0.0), d.get("y", 0.0), d.get("z", 0.0),
            d.get("rotation", 0.0),
            d.get("voltage", 0.0), d.get("current", 0.0), d.get("load", 0.0),
            d.get("createdAt", ""), d.get("updatedAt", ""),
            props_str,
        ]
        for col, val in enumerate(row_vals, start=1):
            ws_dev.cell(row=i, column=col, value=val)
    # Auto-size
    for col_idx in range(1, len(dev_cols) + 1):
        ws_dev.column_dimensions[get_column_letter(col_idx)].width = 16
    ws_dev.column_dimensions["N"].width = 40  # Properties

    # ── Sheet 3: Connections ───────────────────────────────────────
    ws_conn = wb.create_sheet("Connections")
    conn_cols = ["ID", "From ID", "To ID", "Cable Size", "Length (m)", "Type", "Created At"]
    for col, name in enumerate(conn_cols, start=1):
        cell = ws_conn.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, c in enumerate(connections, start=2):
        row_vals = [
            c.get("id", ""), c.get("fromId", ""), c.get("toId", ""),
            c.get("cableSize", ""), c.get("length", 0.0),
            c.get("type", ""), c.get("createdAt", ""),
        ]
        for col, val in enumerate(row_vals, start=1):
            ws_conn.cell(row=i, column=col, value=val)
    for col_idx in range(1, len(conn_cols) + 1):
        ws_conn.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 4: Bill of Quantities (deterministic, no Math.random) ─
    ws_boq = wb.create_sheet("Bill of Quantities")
    boq_cols = ["Category", "Type", "Count", "Unit", "Notes"]
    for col, name in enumerate(boq_cols, start=1):
        cell = ws_boq.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Aggregate devices by (category, type)
    agg: Dict[Tuple[str, str], int] = {}
    for d in devices:
        key = (str(d.get("category", "unknown")), str(d.get("type", "unknown")))
        agg[key] = agg.get(key, 0) + 1
    # Cable aggregation by size
    cable_agg: Dict[str, float] = {}
    for c in connections:
        size = str(c.get("cableSize", "unknown"))
        cable_agg[size] = cable_agg.get(size, 0.0) + float(c.get("length", 0.0))
    row_idx = 2
    for (cat, typ), count in sorted(agg.items()):
        ws_boq.cell(row=row_idx, column=1, value=cat)
        ws_boq.cell(row=row_idx, column=2, value=typ)
        ws_boq.cell(row=row_idx, column=3, value=count)
        ws_boq.cell(row=row_idx, column=4, value="unit")
        ws_boq.cell(row=row_idx, column=5, value="Fire alarm device per NFPA 72")
        row_idx += 1
    # Cable rows
    for size, total_len in sorted(cable_agg.items()):
        ws_boq.cell(row=row_idx, column=1, value="Cable")
        ws_boq.cell(row=row_idx, column=2, value=size)
        ws_boq.cell(row=row_idx, column=3, value=round(total_len, 2))
        ws_boq.cell(row=row_idx, column=4, value="m")
        ws_boq.cell(row=row_idx, column=5, value="Total cable length per NEC Ch. 9")
        row_idx += 1
    for col_idx in range(1, len(boq_cols) + 1):
        ws_boq.column_dimensions[get_column_letter(col_idx)].width = 20
    ws_boq.column_dimensions["E"].width = 40

    # ── Serialize to .xlsx bytes ───────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = f"{_safe_filename(project.get('name', 'project'))}_export.xlsx"
    
    return content, media_type, filename


def _generate_manifest_export(project, devices, connections, export_type, project_id):
    """Generate manifest export content for non-Excel formats."""
    manifest = {
        "project": {
            "id": project.get("id", ""),
            "name": project.get("name", ""),
            "author": project.get("author", ""),
        },
        "exportType": export_type,
        "deviceCount": len(devices),
        "connectionCount": len(connections),
        "availableEndpoints": [
            f"/api/v1/projects/{project_id}/export/dxf",
            f"/api/v1/projects/{project_id}/export/revit",
            f"/api/v1/projects/{project_id}/export/ifc",
            "/api/v1/exports (this endpoint, with exportType=excel)",
        ],
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "software": "FireAI / BAZSPARK v1.55.0 (V213)",
        "note": (
            f"exportType='{export_type}' is not a binary format. Use the "
            "availableEndpoints above to fetch a real DXF/Revit/IFC export, "
            "or use exportType='excel' on this endpoint for a real .xlsx workbook."
        ),
    }
    content = json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")
    media_type = "application/json"
    filename = f"{_safe_filename(project.get('name', 'project'))}_manifest.json"
    
    return content, media_type, filename


class ExportDataInput(BaseModel):
    exportType: str
    dataIds: Optional[list] = None


@project_router.post(
    "",
    status_code=200,
    dependencies=[Depends(require_permission(Permission.EXPORT_READ))],
    responses={
        404: {"description": "No projects found to export data"},
    },
)
@limiter.limit("10/minute")
async def export_data_global(request: Request, input_data: ExportDataInput):  # NOSONAR — S3776
    """
    Export project data globally using the first available project for compatibility.

    V213 FIX (Rule 1 — Truthfulness): Previously this endpoint returned
    ``b"MOCK EXCEL EXPORT DATA"`` (13 bytes of plain text) with a fake
    ``.xlsx`` MIME type — opening the file in Excel would fail. Now it
    produces a real ``.xlsx`` workbook via ``openpyxl`` containing:
      - Sheet "Project": project metadata + export timestamp + software version
      - Sheet "Devices": full device inventory (id, name, type, category,
        coordinates, voltage, current, load, properties)
      - Sheet "Connections": full wiring list (from→to, cable size, length)
      - Sheet "Bill of Quantities": deterministic device/cable counts

    For other export types, returns a real JSON manifest (instead of
    ``b"MOCK EXPORT DATA"``) so the client can see what was exported.
    """
    db = get_db()
    projects = db.list_projects(page=1, limit=1)
    if not projects or not projects.get("data"):
        raise HTTPException(status_code=404, detail="No projects found to export data")  # NOSONAR — S8415

    project_id = projects["data"][0]["id"]
    project = projects["data"][0]
    export_type = input_data.exportType.lower()

    # Pull the same data the real DXF/IFC exports use — no fabrication
    devices = db.get_all_devices_for_project(project_id)
    connections = db.get_all_connections_for_project(project_id)

    if export_type == "excel":
        content, media_type, filename = _generate_excel_export(project, devices, connections)
    else:
        content, media_type, filename = _generate_manifest_export(project, devices, connections, export_type, project_id)

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )